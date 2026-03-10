#!/usr/bin/env python3
"""
Enrich O*NET occupations CSV with education, job titles, descriptions, wages,
growth, and job openings.

Data sources (by priority):
- Education: 1) scraped from O*NET Online survey section, 2) O*NET DB files
- Job descriptions & titles: O*NET DB files, fallback to old scrape cache
- Wages, growth labels, openings: scraped from O*NET Online (not in DB)
- Growth numeric: BLS Employment Projections CSV

Uses a JSON cache for resumability of scraped data.
Outputs to intermediate CSV files in the data directory.
"""

import csv
import json
import re
import time
import urllib.request
from html.parser import HTMLParser
from pathlib import Path

import openpyxl

DATA_DIR = Path(__file__).parent.parent / "data"
INPUT_CSV = DATA_DIR / "input" / "All_Occupations_ONET.csv"
EMPLOYMENT_PROJECTIONS_CSV = DATA_DIR / "input" / "Employment Projections.csv"
ONET_DB_DIR = DATA_DIR / "input" / "onet_db"
CACHE_FILE = DATA_DIR / "intermediate" / "onet_scrape_cache.json"
ENRICHMENT_ONLY_CSV = DATA_DIR / "intermediate" / "onet_enrichment.csv"
ENRICHED_CSV = DATA_DIR / "intermediate" / "All_Occupations_ONET_enriched.csv"

HEADERS = {"User-Agent": "Mozilla/5.0 (research project)"}
DELAY = 1.0  # seconds between requests to be polite

# Map ETE category numbers to short labels matching our existing output format.
# Category descriptions from ETE Categories.xlsx, shortened for CSV output.
EDUCATION_CATEGORY_LABELS = {
    1: "Less than high school diploma",
    2: "High school diploma or equivalent",
    3: "Post-secondary certificate",
    4: "Some college, no degree",
    5: "Associate's degree",
    6: "Bachelor's degree",
    7: "Post-baccalaureate certificate",
    8: "Master's degree",
    9: "Post-master's certificate",
    10: "Professional degree",
    11: "Doctoral degree",
    12: "Post-doctoral training",
}


# ---------------------------------------------------------------------------
# O*NET Database file loaders (no scraping needed)
# ---------------------------------------------------------------------------

def load_occupation_data() -> dict:
    """Load job descriptions from Occupation Data.xlsx.
    Returns {code: description}."""
    path = ONET_DB_DIR / "Occupation Data.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, _title, description = row[0], row[1], row[2]
        if code and description:
            lookup[code] = description
    wb.close()
    return lookup


def load_sample_titles() -> dict:
    """Load sample job titles from Sample of Reported Titles.xlsx.
    Returns {code: "Title1, Title2, ..."}."""
    path = ONET_DB_DIR / "Sample of Reported Titles.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    by_code = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, _title, reported_title, shown = row[0], row[1], row[2], row[3]
        if code and reported_title:
            by_code.setdefault(code, []).append(reported_title)
    wb.close()
    return {code: ", ".join(titles) for code, titles in by_code.items()}


def load_education_data() -> dict:
    """Load education levels from Education Training and Experience.xlsx.
    Returns {code: {"education_top_2": "...", "top_education_level": "...", "top_education_rate": "..."}}."""
    path = ONET_DB_DIR / "Education Training and Experience.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    # Collect education percentages per occupation
    # Filter: Element Name == "Required Level of Education", Scale ID == "RL"
    by_code = {}  # {code: [(category, data_value), ...]}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        element_name = row[3]
        scale_id = row[4]
        category = row[6]
        data_value = row[7]
        if element_name == "Required Level of Education" and scale_id == "RL":
            if code and category and data_value is not None and data_value > 0:
                by_code.setdefault(code, []).append((category, data_value))
    wb.close()

    result = {}
    for code, items in by_code.items():
        # Sort by percentage descending, take top 2
        items.sort(key=lambda x: x[1], reverse=True)
        top_2 = items[:2]

        # Format: "63% High school diploma or equivalent | 20% Some college, no degree"
        parts = []
        for cat, val in top_2:
            label = EDUCATION_CATEGORY_LABELS.get(cat, f"Category {cat}")
            pct = int(round(val))
            parts.append(f"{pct}% {label}")

        education_str = " | ".join(parts)

        # Top education = first entry
        top_cat, top_val = top_2[0]
        top_level = EDUCATION_CATEGORY_LABELS.get(top_cat, "")
        top_rate = f"{int(round(top_val))}%"

        result[code] = {
            "education_top_2": education_str,
            "top_education_level": top_level,
            "top_education_rate": top_rate,
        }
    return result


# ---------------------------------------------------------------------------
# Employment projections (already downloaded CSV)
# ---------------------------------------------------------------------------

def load_employment_projections() -> dict:
    """Load Employment Projections.csv → {code: percent_change_str}."""
    growth_lookup = {}
    if not EMPLOYMENT_PROJECTIONS_CSV.exists():
        print(f"Warning: {EMPLOYMENT_PROJECTIONS_CSV} not found. Skipping growth number enrichment.")
        return growth_lookup

    with open(EMPLOYMENT_PROJECTIONS_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_code = row.get("Occupation Code", "").strip()
            if raw_code.startswith('="') and raw_code.endswith('"'):
                code = raw_code[2:-1]
            else:
                code = raw_code
            growth_str = row.get("Employment Percent Change, 2024-2034", "").strip()
            if code and growth_str:
                growth_lookup[code] = growth_str
    return growth_lookup


# ---------------------------------------------------------------------------
# Web scraper — wages, growth, openings, and education survey responses
# ---------------------------------------------------------------------------

class OnetPageParser(HTMLParser):
    """Extract wage, growth, openings, and education survey from an O*NET page."""

    def __init__(self):
        super().__init__()
        self._in_dt = False
        self._in_dd = False
        self._current_field = None
        self._capture = False
        self._text_buf = []

        # Education survey list items
        self._in_education_list = False
        self._education_items = []

        self.median_wage = None
        self.projected_growth = None
        self.projected_job_openings = None
        self.education_top_2 = None
        self.jobzone_education_text = None  # prose fallback from Job Zone table

    def handle_starttag(self, tag, attrs):
        if tag == "dt":
            self._in_dt = True
            self._text_buf = []
        elif tag == "dd" and self._current_field:
            self._in_dd = True
            self._capture = True
            self._text_buf = []
        elif tag == "li" and self._in_education_list:
            self._text_buf = []

    def handle_endtag(self, tag):
        if tag == "dt":
            self._in_dt = False
            dt_text = "".join(self._text_buf).strip()
            if "Median wages" in dt_text:
                self._current_field = "wage"
            elif "Projected growth" in dt_text:
                self._current_field = "growth"
            elif "Projected job openings" in dt_text:
                self._current_field = "openings"
            elif dt_text == "Education":
                self._current_field = "jobzone_education"
            else:
                self._current_field = None
        elif tag == "dd" and self._capture:
            self._in_dd = False
            self._capture = False
            dd_text = re.sub(r"\s+", " ", "".join(self._text_buf).strip())
            if self._current_field == "wage":
                self.median_wage = dd_text
            elif self._current_field == "growth":
                self.projected_growth = dd_text
            elif self._current_field == "openings":
                self.projected_job_openings = dd_text
            elif self._current_field == "jobzone_education":
                self.jobzone_education_text = dd_text
            self._current_field = None
        elif tag == "li" and self._in_education_list:
            li_text = "".join(self._text_buf).strip()
            if li_text and "%" in li_text:
                self._education_items.append(li_text)

    def handle_data(self, data):
        if self._in_dt or self._capture:
            self._text_buf.append(data)
        elif self._in_education_list:
            self._text_buf.append(data)

        # Detect education survey section
        if "How much education does a new hire need" in data or "Respondents said" in data:
            self._in_education_list = True

    def finalize_education(self):
        """Extract top 2 education responses after parsing is complete."""
        if not self._education_items:
            return

        cleaned = []
        for item in self._education_items:
            text = re.sub(r"\s+", " ", item).strip()
            text = text.replace("requiredmore info", "")
            text = text.replace("required", "").strip()
            text = text.replace("responded:", "").strip()
            match = re.search(r"(\d+)\s*%\s*(.*)", text)
            if match:
                pct = match.group(1)
                level = match.group(2).strip()
                if level:
                    cleaned.append((int(pct), f"{pct}% {level}"))

        if cleaned:
            cleaned.sort(key=lambda x: x[0], reverse=True)
            self.education_top_2 = " | ".join(item[1] for item in cleaned[:2])


def _education_from_jobzone(text: str) -> str:
    """Map Job Zone Education prose to a bare education level label.

    Example input: "Most of these occupations require a four-year bachelor's degree, but some do not."
    Example output: "Bachelor's degree"
    """
    t = text.replace("\u2018", "'").replace("\u2019", "'").lower()
    patterns = [
        ("post-doctoral", "Post-doctoral training"),
        ("doctoral degree", "Doctoral degree"),
        ("phd", "Doctoral degree"),
        ("doctor of", "Doctoral degree"),
        ("master's degree", "Master's degree"),
        ("post-baccalaureate", "Post-baccalaureate certificate"),
        ("bachelor's degree", "Bachelor's degree"),
        ("associate's degree", "Associate's degree"),
        ("post-secondary certificate", "Post-secondary certificate"),
        ("vocational", "Post-secondary certificate"),
        ("high school diploma", "High school diploma or equivalent"),
        ("high school", "High school diploma or equivalent"),
        ("less than high school", "Less than high school diploma"),
    ]
    for pattern, label in patterns:
        if pattern in t:
            return label
    return ""


def fetch_onet_page(url: str) -> dict:
    """Scrape wage, growth, openings, and education from an O*NET page."""
    req = urllib.request.Request(url, headers=HEADERS)
    resp = urllib.request.urlopen(req, timeout=30)
    html = resp.read().decode("utf-8", errors="replace")

    parser = OnetPageParser()
    parser.feed(html)
    parser.finalize_education()

    return {
        "median_wage": parser.median_wage or "",
        "projected_growth": parser.projected_growth or "",
        "projected_job_openings": parser.projected_job_openings or "",
        "education_top_2": parser.education_top_2 or "",
        "jobzone_education_text": parser.jobzone_education_text or "",
    }


# ---------------------------------------------------------------------------
# CSV output helpers
# ---------------------------------------------------------------------------

def extract_top_education(education_str: str) -> tuple:
    """Extract top education level and rate from education string.

    Handles two formats:
    1. Standard: "63% High school diploma | 20% Some college" → ("High school diploma", "63%")
    2. Alternate (no %): "Bachelor's degree" → ("Bachelor's degree", "")
    """
    if not education_str:
        return ("", "")
    education_str = education_str.replace("\u2018", "'").replace("\u2019", "'").strip()
    if not education_str:
        return ("", "")

    first_item = education_str.split("|")[0].strip() if "|" in education_str else education_str
    match = re.match(r"(\d+%)\s+(.*)", first_item)
    if match:
        return (match.group(2).strip(), match.group(1))
    return (education_str, "")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ENRICHED_CSV.parent.mkdir(parents=True, exist_ok=True)

    # 1. Load O*NET database files (instant, no scraping)
    print("Loading O*NET database files...")
    descriptions = load_occupation_data()
    print(f"  Occupation Data: {len(descriptions)} descriptions")
    titles = load_sample_titles()
    print(f"  Sample Titles: {len(titles)} occupations")
    db_education = load_education_data()
    print(f"  Education (DB fallback): {len(db_education)} occupations")

    # 2. Load BLS employment projections
    growth_lookup = load_employment_projections()
    print(f"  Employment Projections: {len(growth_lookup)} occupations")

    # 3. Load scrape cache
    scrape_cache = {}
    if CACHE_FILE.exists():
        with open(CACHE_FILE) as f:
            scrape_cache = json.load(f)

    # Also migrate old cache format if it exists
    old_cache = DATA_DIR / "intermediate" / "onet_enrichment_cache.json"
    old_cache_data = {}
    if old_cache.exists():
        with open(old_cache) as f:
            old_cache_data = json.load(f)
        # Migrate into scrape cache (wage/growth/openings + education)
        if not CACHE_FILE.exists():
            for code, entry in old_cache_data.items():
                if code not in scrape_cache:
                    scrape_cache[code] = {
                        "median_wage": entry.get("median_wage", ""),
                        "projected_growth": entry.get("projected_growth", ""),
                        "projected_job_openings": entry.get("projected_job_openings", ""),
                        "education_top_2": entry.get("education_top_2", ""),
                    }
            print(f"  Migrated {len(scrape_cache)} entries from old cache")

    # 4. Read input CSV
    with open(INPUT_CSV, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    total = len(rows)
    cached_count = sum(1 for r in rows if r["Code"] in scrape_cache
                       and scrape_cache[r["Code"]].get("median_wage"))
    print(f"\nTotal occupations: {total}, cached: {cached_count}")

    # 5. Scrape O*NET pages for occupations not yet cached
    to_scrape = [r for r in rows
                 if r["Code"] not in scrape_cache
                 or not scrape_cache[r["Code"]].get("median_wage")]
    if to_scrape:
        print(f"Scraping {len(to_scrape)} occupations...")
        for i, row in enumerate(to_scrape):
            code = row["Code"]
            url = row["url"]
            print(f"  [{i+1}/{len(to_scrape)}] {code} - {row['Occupation']}...")
            try:
                data = fetch_onet_page(url)
                scrape_cache[code] = data
                print(f"    Wage: {data['median_wage']}, Edu: {data['education_top_2'][:50]}")
            except Exception as e:
                print(f"    ERROR: {e}")
                scrape_cache[code] = {
                    "median_wage": "",
                    "projected_growth": "",
                    "projected_job_openings": "",
                    "education_top_2": "",
                    "jobzone_education_text": "",
                }

            with open(CACHE_FILE, "w") as f:
                json.dump(scrape_cache, f, indent=2)
            time.sleep(DELAY)
    else:
        print("All pages already cached.")

    # 6. Build enrichment and write CSVs
    enrichment_fieldnames = ["Code", "Median Wage", "Projected Growth",
                             "Employment Change, 2024-2034", "Projected Job Openings",
                             "Education", "Top Education Level", "Top Education Rate",
                             "Sample Job Titles", "Job Description"]

    # Report codes with gaps in DB files
    all_input_codes = {r["Code"] for r in rows}
    missing_desc = all_input_codes - set(descriptions.keys())
    missing_titles = all_input_codes - set(titles.keys())
    any_missing = missing_desc | missing_titles
    if any_missing:
        print(f"\nCodes with gaps in O*NET DB files (SOC version mismatch):")
        print(f"  Missing descriptions: {len(missing_desc)}, titles: {len(missing_titles)}")
        if old_cache_data:
            fallback_count = sum(1 for c in any_missing if c in old_cache_data)
            print(f"  Using old scrape cache as fallback for {fallback_count} of them.")

    def build_enrichment(code):
        scraped = scrape_cache.get(code, {})
        desc = descriptions.get(code, "")
        job_titles = titles.get(code, "")

        # Education priority:
        #   1) Scraped survey <li> items (percentages from respondents)
        #   2) O*NET DB file (structured survey data, may lag website)
        #   3) Job Zone Education prose (bare label, no percentage)
        edu_str = scraped.get("education_top_2", "")
        if edu_str:
            top_level, top_rate = extract_top_education(edu_str)
        else:
            db_edu = db_education.get(code, {})
            edu_str = db_edu.get("education_top_2", "")
            top_level = db_edu.get("top_education_level", "")
            top_rate = db_edu.get("top_education_rate", "")
            if not edu_str:
                # Fallback: Job Zone Education prose → bare label, no percentage
                jz_text = scraped.get("jobzone_education_text", "")
                top_level = _education_from_jobzone(jz_text) if jz_text else ""
                edu_str = top_level
                top_rate = ""

        # Fallback per field: if missing from DB, use old scrape cache
        if code in old_cache_data:
            old = old_cache_data[code]
            if not desc:
                desc = old.get("job_description", "")
            if not job_titles:
                job_titles = old.get("sample_job_titles", "")

        percent_change = growth_lookup.get(code, "")
        if not percent_change and code.endswith(".00"):
            percent_change = growth_lookup.get(code[:-3], "")
        return {
            "Median Wage": scraped.get("median_wage", ""),
            "Projected Growth": scraped.get("projected_growth", ""),
            "Employment Change, 2024-2034": percent_change,
            "Projected Job Openings": scraped.get("projected_job_openings", ""),
            "Education": edu_str,
            "Top Education Level": top_level,
            "Top Education Rate": top_rate,
            "Sample Job Titles": job_titles,
            "Job Description": desc,
        }

    with open(ENRICHMENT_ONLY_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=enrichment_fieldnames)
        writer.writeheader()
        for row in rows:
            enrichment = build_enrichment(row["Code"])
            enrichment["Code"] = row["Code"]
            writer.writerow(enrichment)

    # 7. Write fully enriched CSV
    enrichment_cols = ["Median Wage", "Projected Growth", "Employment Change, 2024-2034",
                       "Projected Job Openings", "Education", "Top Education Level",
                       "Top Education Rate", "Sample Job Titles", "Job Description"]
    existing = list(rows[0].keys())
    fieldnames = existing + [c for c in enrichment_cols if c not in existing]

    with open(ENRICHED_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            enrichment = build_enrichment(row["Code"])
            row.update(enrichment)
            writer.writerow(row)

    # Report education source stats
    scraped_edu = sum(1 for r in rows if scrape_cache.get(r["Code"], {}).get("education_top_2"))
    db_edu_used = sum(1 for r in rows
                      if not scrape_cache.get(r["Code"], {}).get("education_top_2")
                      and r["Code"] in db_education)
    no_edu = total - scraped_edu - db_edu_used
    print(f"\nEducation sources: {scraped_edu} scraped, {db_edu_used} DB fallback, {no_edu} none")

    print(f"\nDone!")
    print(f"  Enrichment fields only: {ENRICHMENT_ONLY_CSV}")
    print(f"  Fully enriched CSV: {ENRICHED_CSV}")


if __name__ == "__main__":
    main()
